__author__ = 'syan'
import os
import re
import sys
import inspect

from pyExcelerator import *

DEBUG = 0


def get_this_file():
    this_file = inspect.getfile(inspect.currentframe())
    return os.path.abspath(this_file)
this_dir = os.path.join(os.path.dirname(get_this_file()), "xlib")
# --------------------------------
_cellBorderLineStyle = {'0' : 0x00,   # NO_LINE
                        #'1' : 0x01,   # THIN
                        '1' : 0x03,    # remove line
                        '2' : 0x02,   # MEDIUM
                        '3' : 0x03,   # DASHED
                        '4' : 0x04,   # DOTTED
                        '5' : 0x05,   # THICK
                        '6' : 0x06,   # DOUBLE
                        '7' : 0x07}   # HAIR
def createBorders(lineStyle='0000'):
    myBorders = Borders()
    lineStyleList = list(lineStyle)
    myBorders.top    = _cellBorderLineStyle[lineStyleList[0]]
    myBorders.bottom = _cellBorderLineStyle[lineStyleList[1]]
    myBorders.left   = _cellBorderLineStyle[lineStyleList[2]]
    myBorders.right  = _cellBorderLineStyle[lineStyleList[3]]
    return myBorders
_cellAlignmentStyle = {'100' : 0x01,  # left
                       '010' : 0x02,  # center
                       '001' : 0x03}  # right
def createAlignment(cellAlignment='100'):
    myAlignment = Alignment()
    myAlignment.horz = _cellAlignmentStyle[cellAlignment]
    myAlignment.vert = 0x01  # VERTICAL CENTER
    return myAlignment
def createPattern(cellPatternColour=''):
    myPattern = Pattern()
    if cellPatternColour:
        myPattern.pattern = 0x01
        myPattern.pattern_back_colour = cellPatternColour
        myPattern.pattern_fore_colour = cellPatternColour
    return myPattern
def createFont(cellFontColour='black', cellFontIsBold=False, cellFontSize=8):
    myFont = Font()
    myFont.colour_index = cellFontColour
    myFont.bold = cellFontIsBold
    myFont.height = cellFontSize * 20   # IN PIXEL
    return myFont
def createStyle(myBorders, myAlignment, myPattern, myFont, myFormat=''):
    myStyle = XFStyle()
    myStyle.borders = myBorders
    myStyle.alignment = myAlignment
    myStyle.pattern = myPattern
    myStyle.font = myFont
    if myFormat:
        myStyle.num_format_str = myFormat
    return myStyle

myBorders, myAlignment, myPattern, myFont= createBorders(), createAlignment(), createPattern(), createFont()
def_style = createStyle(myBorders, myAlignment, myPattern, myFont, myFormat='')
red_style = createStyle(myBorders, myAlignment, myPattern, createFont("red", cellFontIsBold=True), myFormat='')
green_style = createStyle(myBorders, myAlignment, createPattern(170), createFont("black", cellFontIsBold=True), myFormat='')
yellow_style = createStyle(myBorders, myAlignment, createPattern(171), createFont("black", cellFontIsBold=True), myFormat='')
title_style = createStyle(myBorders, createAlignment('010'), myPattern, createFont(cellFontIsBold=True, cellFontSize=20))
head2_style = createStyle(myBorders, createAlignment('010'), myPattern, createFont(cellFontIsBold=True))
# --------------------------------
def xls2csv(xls_file):
    matrix = [[]]
    sheet_1st = ""
    for sheet_name, values in parse_xls(xls_file, 'gb2312'): # parse_xls(arg) -- default encoding
        sheet_1st = sheet_name.encode('gb2312', 'backslashreplace')
        if DEBUG: print 'Sheet = "%s"' % sheet_1st
        for row_idx, col_idx in sorted(values.keys()):
            v = values[(row_idx, col_idx)]
            if isinstance(v, unicode):
                v = v.encode('gb2312', 'backslashreplace')
            else:
                v = str(v)
            last_row, last_col = len(matrix), len(matrix[-1])
            while last_row < row_idx:
                matrix.extend([[]])
                last_row = len(matrix)

            while last_col < col_idx:
                matrix[-1].extend([''])
                last_col = len(matrix[-1])

            matrix[-1].extend([v])
        break # only transfer the first sheet
    return sheet_1st, matrix

from openpyxl.reader.excel import load_workbook
def xlsx2csv(xlsx_file):
    matrix = [[]]
    sheet_1st = ""
    work_book = load_workbook(filename=xlsx_file)
    if DEBUG: print work_book.get_sheet_names()
    for sheet_name in work_book.get_sheet_names():
        sheet_1st = sheet_name.encode('gb2312', 'backslashreplace')
        if DEBUG: print 'Sheet = "%s"' % sheet_1st
        work_sheet = work_book.get_sheet_by_name(sheet_name)
        col_mx = work_sheet.get_highest_column()
        row_mx = work_sheet.get_highest_row()
        for rx in range(row_mx):
            col_content = list()
            for cx in range(col_mx):
                t_value = work_sheet.cell(row=rx, column=cx).value
                if not t_value:
                    t_value = ""
                else:
                    try:
                        t_value = t_value.encode('gb2312').strip()
                    except AttributeError:
                        t_value = str(t_value)
                    except UnicodeEncodeError:
                        t_value = ""
                col_content.append(t_value)
            matrix.append(col_content)
        break # only transfer the first sheet
    return sheet_1st, matrix

def excel2csv(excel_file):
    t_file = excel_file.lower()
    sheet_1st = ""
    csv = list()
    if t_file.endswith(".xls"):
        sheet_1st, csv = xls2csv(excel_file)
    elif t_file.endswith(".xlsx"):
        sheet_1st, csv = xlsx2csv(excel_file)
    else:
        print "can not transfer to csv format for %s. xls/xlsx file needed." % excel_file
    return sheet_1st, csv

def check_cr_title(current_title):
    """
    must be same as standard order
    """
    print "-----", current_title
    standard_title = ["Key", "Summary", "Priority", "Status", "Entry Type"]
    standard_title = [item.lower() for item in standard_title]
    standard_title = set(standard_title)
    set_cur_title = [item.lower() for item in current_title]
    set_cur_title = set(set_cur_title)
    diff_title = standard_title - set_cur_title
    if diff_title:
        print("Warning. Not found standard CR title: %s" % ",".join(current_title))


def get_cr_content(csv_list):
    cr_title, cr_content = list(), list()
    status_idx = 0 # modified.  now is priority index name
    pv_idx = 0 # Product Version index
    real_sts_idx = 0  # real status index
    for item in csv_list:
        if not item: continue
        if item[0].upper() in ("ID", "KEY"):
            cr_title = [bar.lower() for bar in item]
            try:
                status_idx = cr_title.index("priority")
                real_sts_idx = cr_title.index("status")
                pv_idx = cr_title.index("product version")
            except ValueError, e:
                print "Warning", e
            continue
        if not real_sts_idx:
            continue

        try:
            real_sts = item[real_sts_idx]
            if real_sts.lower() in ("duplicate", "enhancement", "nab"):
                continue
            if status_idx:
                cr_content.append(item)
        except Exception, e:
            if DEBUG:
                print e
            pass

    if DEBUG:
        print cr_title
        print cr_content
    return cr_title, cr_content, status_idx, pv_idx

def create_xls_file(xls_file, xls_content, diamond_version):
    wb = Workbook()
    ws = wb.add_sheet("Summary_INDEX")
    n = "HYPERLINK"
    ws.col(8).width = ws.col(9).width = 200*20
    for i, item in enumerate(xls_content):
        if len(item) == 1: # title
            ws.write_merge(0,0, 0, 8, item[0], title_style)
            ws.write_merge(1,1, 1, 7, "SourceReport", head2_style)
            ws.write(1, 0, "#", head2_style)
            ws.write(1, 8, "TotalCRNum", head2_style)
            ws.write(1, 9, "BlockerNum", head2_style)
            ws.write(1, 10, "CRsBefore%s" % diamond_version, head2_style)
        else:
            color, sheet_name, ori_file_name, all_cr_no, blocker_no, cr_before_no = item
            if color == "RED":
                style = red_style
            elif color == "GREEN":
                style = green_style
            elif color == "YELLOW":
                style = yellow_style
            else:
                style = def_style
            ws.write_merge(i+1, i+1, 1, 7, Formula(n + '("#\'%s\'!A2";"%s")' % (sheet_name, ori_file_name)), style)
            ws.write(i+1, 0, i, style)
            ws.write(i+1, 8, all_cr_no, style)
            ws.write(i+1, 9, blocker_no, style)
            ws.write(i+1, 10, cr_before_no, style)
    wb.save(xls_file)

# SOF-102847 ISPL_CR_58035
p_cr_key = re.compile("(sof|CR)\S\d+", re.I)
def get_tab_color(cr_content, sts_idx , pv_idx, diamond_version):
    """
    RED: have blocker CR
    YELLOW: have CR except blocker
    GREEN: No CR
    """
    color = "GREEN"
    all_cr_no, blocker_no = 0, 0
    cr_before = 0
    t_dv = re.sub("\.", "\.", diamond_version)
    p_cr_before = re.compile(t_dv)
    for content in cr_content:
        if not p_cr_key.search(content[0]):  # not a CR number
            continue
        if color != "RED":
            color = "YELLOW"   # specify YELLOW
        all_cr_no += 1
        content = [item.lower() for item in content]
        sts = content[sts_idx]
        if not pv_idx:
            cr_before = "No Version"
        else:
            try:
                pv = content[pv_idx]
            except IndexError:
                pv = "NA"
            if not p_cr_before.search(pv):
                cr_before += 1
        if sts in ("blocker", "pri 0", "pri0"):
            blocker_no += 1
            color = "RED"

    return color, all_cr_no, blocker_no, cr_before

p_diamond = re.compile("(diamond\S+)\s+(.+)\s+TestReport\.", re.I)

def create_new_file_list(file_list, diamond_version, idx_sheet_title=""):
    index_xls, index_content = os.path.abspath("DiamondSummary Index TestReport.xls"), list()
    new_file_list = list()
    new_file_list.append('GREEN@%s' % index_xls)
    for x_file in open(file_list):
        # ----------------------------------
        # skip the lines as following.
        x_file = x_file.strip()
        if not x_file: continue
        if x_file.startswith("#"): continue
        if x_file.startswith(";"): continue
        t_x_file = os.path.basename(x_file)
        print "parsing %s" % t_x_file
        m_diamond = p_diamond.search(t_x_file)
        # ----------------------------------
        # get Diamond version
        if not index_content:
            # only find Diamond Version when index_content is empty
            if idx_sheet_title:
                index_content.append([idx_sheet_title])
            else:
                if not m_diamond:
                    # must find diamond version
                    print("Error. Not found Diamond version from %s" % x_file)
                    continue
                index_content.append([m_diamond.group(1)+" Reports Index Page"])
                # get like: Diamond3.3.0.78 Reports Index Page

        sheet_1st, csv_list = excel2csv(x_file)
        if not sheet_1st:
            print("Error. No sheet found in %s" % x_file)
            continue
        cr_title, cr_content, sts_idx , pv_idx = get_cr_content(csv_list)
        # ----------------------------------
        # set TAB color, get index content
        if not cr_title:
            print("Warning. No CR titles found in %s" % x_file)
        check_cr_title(cr_title)
        # TBD
        tab_color, all_cr_no, blocker_no, cr_before_no = get_tab_color(cr_content, sts_idx , pv_idx, diamond_version)

        index_content.append([tab_color, m_diamond.group(2), t_x_file, all_cr_no, blocker_no, cr_before_no])
        new_file_list.append("%s@%s" % (tab_color, x_file))
    # --------------
    # create index xls file
    create_xls_file(index_xls, index_content, diamond_version)
    # --------------
    # create new summary report file
    new_file = os.path.abspath("t_file_list")
    new_ob = open(new_file, "w")
    for item in new_file_list:
        print >> new_ob, item
    new_ob.close()

    merge_cmd = r"%s %s/MergeExcel.py -f %s -l Excel_Macro.log -r finalReport" % (sys.executable, this_dir, new_file)
    from xlib import xCommand
    sts = xCommand.run_command(merge_cmd, "run.time", "run.log")
    return sts


if __name__ == "__main__":
    import optparse
    parser = optparse.OptionParser()
    parser.add_option("-V", "--diamond", help="specify current Diamond version number, like 3.3, 3.4")
    opts, args = parser.parse_args()
    diamond_version = opts.diamond
    if args:
        file_list = args[0]
    else:
        file_list = "report.list"
    create_new_file_list(file_list, diamond_version)
