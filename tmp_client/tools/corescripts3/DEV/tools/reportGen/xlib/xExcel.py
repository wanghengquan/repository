#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Initial:
  date: 9:09 AM 10/25/2021
  file: xExcel.py
  author: Shawn Yan
Description:

"""
import os
import re
import itertools
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from openpyxl import load_workbook
import xlrd2


def set_cell_value(raw_data, for_csv=False):
    try:
        x = int(raw_data)
    except:
        try:
            x = float(raw_data)
        except:
            x = raw_data
    if for_csv:
        x = str(x)
        if "," in x:
            x = '"{}"'.format(x)
    return x


class BasicExcel(object):
    def __init__(self):
        self.borders_lrtb = self._create_borders_lrtb()
        self.red_borders_lrtb = self._create_borders_lrtb("FF0000")
        self.c_r = "cell_range"
        self.m_r = "merge_range"
        self.wb = Workbook()
        self.number_percentage = "0.00%"
        self.number_float = "0.000"
        self.alignment_left = Alignment(horizontal="left")
        self.alignment_right = Alignment(horizontal="right")
        self.alignment_center = Alignment(horizontal="center")
        self.alignment_center_wrap = Alignment(horizontal="center", wrap_text=True)
        self.font_default = Font(size=9)
        self.font_bold = Font(bold=True, size=9)
        self.font_red = Font(bold=True, size=9, color="FF0000")
        self.font_green = Font(bold=True, size=9, color="00962D")
        self.font_link = Font(bold=True, size=10, color="0000FF", underline="single")
        self.font_link_content = Font(bold=True, size=10, color="0000FF")
        self.font_link_home = Font(bold=True, name="Georgia", size=10, color="0000FF", underline="single")
        self.fill_skip = PatternFill(fill_type='solid', start_color="DAEEF3", end_color="DAEEF3")
        self.fill_out_of_range = PatternFill(fill_type='solid', start_color="F0FFC3", end_color="F0FFC3")

    def save_to_excel(self, excel_file):
        for default_sheet_name in ("Sheet", "Sheet1"):
            try:
                ws = self.wb[default_sheet_name]
                self.wb.remove(ws)
            except KeyError:
                pass
        self.wb.save(excel_file)

    @staticmethod
    def _create_borders_lrtb(color="000000"):
        l_r_t_b = ["left", "right", "top", "bottom"]
        side0, side1 = Side(style="hair", color=color), Side(style="medium", color=color)
        side_dict = {"0": side0, "1": side1, "x": Side()}
        borders = dict()
        for border_keys in itertools.product(list(side_dict.keys()), repeat=4):
            borders["".join(border_keys)] = Border(**dict(zip(l_r_t_b, [side_dict.get(foo) for foo in border_keys])))
        return borders

    @staticmethod
    def set_column_width(ws, column, width=40):
        ws.column_dimensions[get_column_letter(column)].width = width

    def write_row_cells(self, ws, row_data):
        """
        :param ws: worksheet
        :param row_data: cell attributes dictionary list
        cell attribute keys:
           cell_range
           merge_range
           value
           font
           alignment
           border
           number_format
           ...
        """
        for foo in row_data:
            if isinstance(foo, int):
                continue
            cell_range = foo.get(self.c_r)
            merge_range = foo.get(self.m_r)
            if cell_range:
                cell = ws.cell(**cell_range)
            else:
                cell = ws.cell(row=merge_range.get("start_row"), column=merge_range.get("start_column"))
            for k, v in list(foo.items()):
                if k in (self.c_r, self.m_r, "real_value", "real_string"):
                    continue
                if k == "column_width":
                    v = int(v)  # original v is string
                    self.set_column_width(ws, cell_range.get("column"), v)
                else:
                    cell.__setattr__(k, v)
            if merge_range:
                ws.merge_cells(**merge_range)

    def add_sheets_only(self, sheet_name_list):
        for foo in sheet_name_list:
            assert len(foo) < 31, "sheet name {} is too long".format(foo)
        ws = self.wb.active
        ws.title = sheet_name_list[0]
        for sn in sheet_name_list[1:]:
            self.wb.create_sheet(sn)


def none2blank(raw_value):
    if raw_value is None:
        raw_value = ""
    raw_value = str(raw_value).strip()
    try:
        _ = float(raw_value)
    except ValueError:
        raw_value = '"{}"'.format(raw_value)
    return raw_value


class Excel2csv(object):
    def __init__(self, max_row_size=1000, max_column_size=300):
        self.max_row_size = max_row_size
        self.max_column_size = max_column_size
        self._csv_dict = dict()

    def get_csv_dict_alike(self):
        this_dict = dict()
        for k, v in list(self._csv_dict.items()):
            keys = v[0]
            this_dict[k] = list()
            for foo in v[1:]:
                if foo:
                    this_dict[k].append(dict(zip(keys, foo)))
        return this_dict

    def dump_to_file(self, output_directory="."):
        if not os.path.isdir(output_directory):
            os.makedirs(output_directory)
        for k, v in list(self._csv_dict.items()):
            filename = os.path.join(output_directory, "Sheet_{}.csv".format(k))
            with open(filename, "w", encoding="utf-8") as ob:
                for foo in v:
                    new_foo = [none2blank(item) for item in foo]
                    print(",".join(new_foo), file=ob, end="\n")

    def process(self, excel_file, sheet_pattern_string=None):
        comments = "Transferring Excel file: {}".format(excel_file)
        if sheet_pattern_string:
            comments += ", sheet name like {}".format(sheet_pattern_string)
        self._csv_dict = dict()
        sheet_pattern = re.compile(sheet_pattern_string) if sheet_pattern_string else None
        if excel_file.endswith(".xlsx"):
            self.transfer_xlsx_file(excel_file, sheet_pattern)
        else:
            self.transfer_xls_file(excel_file, sheet_pattern)

    def transfer_xlsx_file(self, excel_file, sheet_pattern):
        wb = load_workbook(excel_file, data_only=True)
        for this_name in wb.sheetnames:
            if sheet_pattern:
                if not sheet_pattern.search(this_name):
                    continue
            self._csv_dict.setdefault(this_name, list())
            ws = wb[this_name]
            row_range = min(ws.max_row, self.max_row_size) + 1
            column_range = min(ws.max_column, self.max_column_size) + 1
            for r in range(1, row_range):
                _row = list()
                for c in range(1, column_range):
                    _row.append(ws.cell(row=r, column=c).value)
                self._csv_dict[this_name].append(_row)

    def transfer_xls_file(self, excel_file, sheet_pattern):
        wb = xlrd2.open_workbook(excel_file)
        for i, this_name in enumerate(wb.sheet_names()):
            if sheet_pattern:
                if not sheet_pattern.search(this_name):
                    continue
            self._csv_dict.setdefault(this_name, list())
            ws = wb.sheet_by_index(i)
            row_range = min(ws.nrows, self.max_row_size)
            column_range = min(ws.ncols, self.max_column_size)
            for r in range(row_range):
                _row = list()
                for c in range(column_range):
                    _row.append(ws.cell_value(rowx=r, colx=c))
                if _row:
                    self._csv_dict[this_name].append(_row)
