#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Initial:
  date: 3:50 PM 12/1/2021
  file: xReport.py
  author: Shawn Yan
Description:

"""
import os
import copy
import re
import json
from collections import OrderedDict
from capuchin import xTools
from .xOptions import DeployOptions
from . import xUtils
from . import xExcel
from openpyxl.utils import get_column_letter


class DataReport(DeployOptions):
    def __init__(self, opts_ns):
        super(DataReport, self).__init__(opts_ns)
        self.ew = xExcel.BasicExcel()  # Excel Writer
        self.key_skip = "will_skip_this_row_when_calculate"
        self.title_test_id = "xTestID"

    def process(self):
        self.say_info("Try to generate QoR Report ...")
        if self.deploy_options():
            return 1
        self.final_excel_data_dict = OrderedDict()
        self.report_conf = None
        self.prepare_default_options()
        for self.report_conf in self.report_conf_file_list:
            self.excel_data_dict = OrderedDict()
            self.prepare_base_data()
            self.prepare_standard_base_data()
            self.prepare_row_data()
            self.prepare_summary_data()
        self.printout_excel_file()

    def prepare_default_options(self):
        self.y_general = self.default_conf_options.get("report_general")
        self.y_detail = self.default_conf_options.get("report_detail")
        self.cases_group_dict = dict()
        _p_cases_group_name = re.compile(r"group\s*:\s*(.+)")
        for k, v in list(self.default_conf_options.items()):
            _m = _p_cases_group_name.search(k)
            if _m:
                cg = _m.group(1)
                cg = cg.strip()
                _keys = list(v.keys())
                _keys.sort()
                self.cases_group_dict[cg] = [v.get(foo) for foo in _keys]

    def prepare_base_data(self):
        """
        parse report conf file or command arguments and try to merge with default settings
        """
        self.arg_dict = copy.deepcopy(vars(self.opts))
        if self.report_conf:
            p_sheet = re.compile(r"sheet\s+(.+)", re.I)
            sts, report_options = xTools.get_options(self.report_conf)
            if sts:
                self.say_error("Failed to parse {}".format(self.report_conf))
                self.say_traceback()
                xTools.eexit()
            temp_report_data = dict()
            for k, v in list(report_options.items()):
                m = p_sheet.search(k)
                if not m:
                    self.say_warning("Unknown section name: {}".format(k))
                else:
                    v_index = v.get("index")
                    if v_index:
                        v_index = int(v_index)
                    else:
                        v_index = 1000000
                    v["sheet_name"] = sn = m.group(1)
                    temp_key = "{:08d}{}".format(v_index, sn)
                    temp_report_data[temp_key] = v
            keys = list(temp_report_data.keys())
            keys.sort(key=str.lower)
            for kk in keys:
                vv = temp_report_data.get(kk)
                self.excel_data_dict[vv.get("sheet_name")] = vv
        else:  # by command line
            self.create_excel_data_dict_from_command_line()
        if self.report_conf:
            self.this_summary_sheet_name = "Summary_{}".format(xTools.get_file_name_only(self.report_conf))
        else:
            self.this_summary_sheet_name = "Summary"
        if self.debug:
            self.say_info("Basic Data:")
            for k in (self.excel_data_dict.keys()):
                v = self.excel_data_dict.get(k)
                print(k, "---", v)

    def create_excel_data_dict_from_command_line(self):
        if not self.src:
            self.say_error("Error. not specified value for 'src'")
            xTools.eexit()
        if not self.dst:
            self.say_error("Error. not specified value for 'dst'")
            xTools.eexit()

        if self.src_section or self.dst_section:
            lens_ss, lens_ds = len(self.src_section), len(self.dst_section)
            if max(lens_ss, lens_ds) > 1:
                assert lens_ss == lens_ds, "Number of src/dst sections must be the same: {} {}".format(lens_ss, lens_ds)
            else:
                if not self.src_section:
                    self.src_section = [""]
                if not self.dst_section:
                    self.dst_section = [""]
            local_src_sections = self.src_section[:]
            local_dst_sections = self.dst_section[:]
        else:
            processed_data = dict()
            raw_src_data = xUtils.get_raw("SRC", self.arg_dict, processed_data, self.force, self.say_error, raw_only=1)
            raw_dst_data = xUtils.get_raw("DST", self.arg_dict, processed_data, self.force, self.say_error, raw_only=1)
            self.force = False  # do not dump again
            raw_src_sections = [foo.get("xSectionName") for foo in raw_src_data]
            raw_dst_sections = [foo.get("xSectionName") for foo in raw_dst_data]
            all_sections = list(set(raw_src_sections) | set(raw_dst_sections))
            local_src_sections = local_dst_sections = all_sections
        self.arg_dict["src_section"] = self.src_section = list()
        self.arg_dict["dst_section"] = self.dst_section = list()
        self.arg_dict["src_name"] = self.src_name if self.src_name else self.src
        self.arg_dict["dst_name"] = self.dst_name if self.dst_name else self.dst
        for pos_idx, (k, v) in enumerate(dict(zip(local_src_sections, local_dst_sections)).items()):
            key1 = "{}_{}".format(self.arg_dict["src_name"], k)
            key2 = "{}_{}".format(self.arg_dict["dst_name"], v)
            sheet_name = xUtils.get_vs_name(pos_idx+1, key1, key2)
            sheet_data = copy.deepcopy(self.arg_dict)
            sheet_data["index"] = pos_idx
            if v:
                sheet_data["dst_section"] = v
                vv = v
            else:
                vv = ""
            if k:
                sheet_data["src_section"] = k
                kk = k
            else:
                kk = ""
            sheet_data["group_name"] = sheet_data["sub_name"] = xUtils.get_vs_name(None, kk, vv)
            if self.group_name:
                sheet_data["group_name"] = self.group_name
            self.excel_data_dict["{}".format(sheet_name)] = sheet_data

    def prepare_standard_base_data(self):
        """
        remove items: index, sheet_name, titles, section
        """
        _dt = self.y_general.get("default_titles")
        _dct = self.y_general.get("default_compare_titles")
        assignment_dict = dict(src_section=dict(keywords=("src_section", "section"), default=""),
                               dst_section=dict(keywords=("dst_section", "section"), default=""),
                               src_titles=dict(keywords=("src_titles", "titles"), default="default_titles"),
                               dst_titles=dict(keywords=("dst_titles", "titles"), default="default_titles"),
                               compare_titles=dict(keywords=("compare_titles",), default="default_compare_titles"),
                               calculate=dict(keywords=("calculate",), default=self.calculate),
                               src_name=dict(keywords=("src_name", "src"), dedault=""),
                               dst_name=dict(keywords=("dst_name", "dst"), dedault=""),
                               )
        processed_data = dict()  # a container for processed data
        for sheet_name, sheet_data in list(self.excel_data_dict.items()):
            _src = self.src if self.src else sheet_data.get("src")
            _dst = self.dst if self.dst else sheet_data.get("dst")
            if not (_src and _dst):
                self.say_error("@E: both src and dst must be specified.")
                xTools.eexit()
            sheet_data["src"], sheet_data["dst"] = _src, _dst
            for k, v in list(assignment_dict.items()):
                new_v = xUtils.get_real_list(self.y_general, self.arg_dict, sheet_data,
                                             v.get("keywords"), v.get("default"))
                if k in ("src_titles", "dst_titles"):
                    if "Design" not in new_v:
                        new_v.insert(0, "Design")
                    if self.add_test_id:
                        if self.title_test_id in new_v:
                            continue
                        else:
                            new_v.insert(1, self.title_test_id)
                if k in ("src_name", "dst_name"):
                    new_v = new_v[0]
                sheet_data[k] = new_v
            sheet_data["default_conf_path"] = self.default_conf_path
            sheet_data["report_conf"] = self.report_conf
            sheet_data["output"] = self.output
            sd = sheet_data["src_data"] = xUtils.get_raw("SRC", sheet_data, processed_data, self.force, self.say_error)
            dd = sheet_data["dst_data"] = xUtils.get_raw("DST", sheet_data, processed_data, self.force, self.say_error)
            x_g = sheet_data.get("case_group_name")
            sheet_data["case_group_name"] = self.case_group_name if self.case_group_name else x_g

            src_design_list = list(sd.keys())
            dst_design_list = list(dd.keys())
            src_design_list.sort()
            dst_design_list.sort()
            sheet_data["design_names"] = xUtils.get_design_names(self.cases_group_dict,
                                                                 sheet_data.get("case_group_name"),
                                                                 src_design_list, dst_design_list)
        self.say_debug(json.dumps(self.excel_data_dict, indent=2), "Second Data:", debug=self.debug)

    def prepare_row_data(self):
        for sheet_name, sheet_data in list(self.excel_data_dict.items()):
            row_data = OrderedDict()
            self.row_number = 1
            row_data["row_1"] = self.add_titles(sheet_data)   # check self.row_number and return cells data
            row_data["row_2"] = self.add_titles(sheet_data)   # check self.row_number and return cells data
            for i, dn in enumerate(sheet_data.get("design_names")):
                new_data = self.add_row_data(i+1, dn, sheet_data)
                row_data["ID_{}".format(i)] = new_data
            sheet_data["formula_range"] = formula_range = self.get_vs_formula_range(row_data)
            sheet_data["failed_cases_number"] = self.temp_failed_cases_number
            sheet_data["skipped_cases_number"] = self.temp_skipped_cases_number
            sheet_data["whole_cases_number"] = self.temp_whole_cases_number
            self.add_statistics_data(row_data, formula_range, sheet_data.get("calculate"))
            row_data["row_{}".format(self.row_number)] = self.add_titles(sheet_data)  # again
            self.add_win_tie_lose_failed_data(row_data, formula_range)
            sheet_data["row_data"] = row_data

    def add_titles(self, sheet_data):
        cell_data = OrderedDict()
        src_titles = sheet_data.get("src_titles")
        dst_titles = sheet_data.get("dst_titles")
        compare_titles = sheet_data.get("compare_titles")
        src_name = sheet_data.get("src_name", "No SRC name")
        dst_name = sheet_data.get("dst_name", "No DST name")
        src_section = sheet_data.get("src_section")
        dst_section = sheet_data.get("dst_section")
        str_src_section = "::{}".format(",".join(src_section)) if src_section else ""
        str_dst_section = "::{}".format(",".join(dst_section)) if dst_section else ""
        src_table_name = sheet_data.get("src_table_name")
        dst_table_name = sheet_data.get("dst_table_name")
        src_table_name = src_table_name if src_table_name else "{}{}".format(src_name, str_src_section)
        dst_table_name = dst_table_name if dst_table_name else "{}{}".format(dst_name, str_dst_section)
        if self.row_number == 1:
            _mr = dict(start_row=self.row_number, end_row=self.row_number, start_column=1, end_column=2)
            cell_data["go_home"] = dict(merge_range=_mr, value="Go Home")
            cell_data["go_home"]["hyperlink"] = "#'{}'!A1".format(self.this_summary_sheet_name)
            cell_data["go_home"]["font"] = self.ew.font_link_home
            # cell_data["blank_1"] = dict(cell_range=dict(row=self.row_number, column=1), value="")
            # cell_data["blank_2"] = dict(cell_range=dict(row=self.row_number, column=2), value="='{}'")
            _s, _e = 3, 3 + len(src_titles) - 1 - 1
            _mr = dict(start_row=self.row_number, end_row=self.row_number, start_column=_s, end_column=_e)
            cell_data["src_name"] = dict(merge_range=_mr, value=src_table_name)
            _s, _e = _e + 1, _e + len(dst_titles) - 1
            _mr = dict(start_row=self.row_number, end_row=self.row_number, start_column=_s, end_column=_e)
            cell_data["dst_name"] = dict(merge_range=_mr, value=dst_table_name)
            if self.no_ad_data:
                _s, _e = _e + 1, _e + len(compare_titles)
            else:
                _s, _e = _e + 1, _e + len(compare_titles) * 2
            _mr = dict(start_row=self.row_number, end_row=self.row_number, start_column=_s, end_column=_e)
            cell_data["vs_name"] = dict(merge_range=_mr, value="{} VS {}".format(src_table_name, dst_table_name))
        else:
            cell_data["ID"] = dict(cell_range=dict(row=self.row_number, column=1), value="ID")
            cell_data["Design"] = dict(cell_range=dict(row=self.row_number, column=2), value="Design")

            def __add_cell(raw_titles, start_column, prefix=""):
                for foo in raw_titles:
                    if foo == "Design":
                        continue
                    new_key = prefix + foo
                    cell_data[new_key] = dict(cell_range=dict(row=self.row_number, column=start_column), value=new_key)
                    start_column += 1
                return start_column

            _c = __add_cell(src_titles, 3, prefix="")
            _c = __add_cell(dst_titles, _c, prefix="#")
            _c = __add_cell(compare_titles, _c, prefix="VS_")
            if not self.no_ad_data:
                _c = __add_cell(compare_titles, _c, prefix="AD_")

        # format
        for k, v in list(cell_data.items()):
            v["border"] = self.ew.borders_lrtb["1111"]
            v["alignment"] = self.ew.alignment_center
            if not v.get("font"):
                v["font"] = self.ew.font_bold
            if self.row_number == 2:
                if k == "ID":
                    cell_width = 5
                elif k == "Design":
                    cell_width = 15
                else:
                    k_len = len(k)
                    if k_len > 10:
                        cell_width = k_len
                    else:
                        cell_width = 0
                if cell_width:  # in PEP8, new dictionary key's value type should be in the previous value type list
                    v["column_width"] = str(cell_width)
        self.row_number += 1  # for next row
        return cell_data

    def prepare_summary_data(self):
        summary_data = OrderedDict()
        self.final_excel_data_dict[self.this_summary_sheet_name] = dict(row_data=summary_data)
        for sheet_name, ready_data in list(self.excel_data_dict.items()):
            self.final_excel_data_dict[sheet_name] = ready_data

        original_summary_data, vendor_tool_versions = xUtils.extract_original_summary_data(self.excel_data_dict)
        value_list = xUtils.get_value_list(original_summary_data)

        version_data = dict()
        version_data["one"] = dict(cell_range=dict(row=2, column=2), value="Vendor Tools:", font=self.ew.font_bold)
        for i, v in enumerate(vendor_tool_versions):
            _mr = dict(start_row=3 + i, end_row=3 + i, start_column=3, end_column=5)
            version_data[str(i)] = dict(merge_range=_mr, value=v, font=self.ew.font_bold)
        summary_data["Version"] = version_data
        self.row_number = 8
        for m, t in enumerate(value_list):
            new_t = dict()
            for i, foo in enumerate(t):
                if not foo:
                    continue
                new_t[i] = dict(cell_range=dict(row=self.row_number, column=i + 2), value=foo,
                                border=self.ew.borders_lrtb["1111"], font=self.ew.font_bold)
                if re.search("([><])", foo):
                    if "change " in foo:
                        merge_range = dict(start_row=self.row_number, end_row=self.row_number,
                                           start_column=i+1, end_column=i+2)
                        new_t[i] = dict(merge_range=merge_range, value=foo,
                                        border=self.ew.borders_lrtb["1111"], font=self.ew.font_bold)
                if "SUM" not in foo:
                    new_t[i]["number_format"] = self.ew.number_percentage
                if foo.startswith("LINK::::"):
                    value_list = re.split("::::", foo)
                    new_t[i]["value"] = value_list[1]
                    new_t[i]["hyperlink"] = "#'{}'!A1".format(value_list[-1])
                    new_t[i]["font"] = self.ew.font_link
                elif foo.startswith("VS_"):
                    value_list = re.split("_", foo)
                    new_t[i]["value"] = "_".join(value_list[1:])
                elif foo.startswith("NEW_AVERAGE"):  # NEW_AVERAGE:::{}:::{}
                    new_t[i]["value"] = self.get_summary_average(foo, i + 2)
            self.row_number += 1
            summary_data[m] = new_t

    def get_summary_average(self, raw_string, now_column):
        raw_list = re.split(":::", raw_string)
        numerator, denominator = raw_list[1], raw_list[2]
        for i in range(1, now_column):
            old = "_{}_".format(i)
            new = "${}{}".format(get_column_letter(now_column - i), self.row_number)
            numerator = re.sub(old, new, numerator)
        return '=IFERROR(({})/({}), "-")'.format(numerator, denominator)

    def printout_excel_file(self):
        sheet_names = list(self.final_excel_data_dict.keys())
        summary_names, detail_names = list(), list()
        for foo in sheet_names:
            if "Summary" in foo:
                summary_names.append(foo)
            else:
                detail_names.append(foo)
        self.ew.add_sheets_only(summary_names + detail_names)
        for sheet_name, ready_data in list(self.final_excel_data_dict.items()):
            this_ws = self.ew.wb[sheet_name]
            if "Summary" not in sheet_name:
                this_ws.freeze_panes = "C3"
            else:
                self.ew.set_column_width(this_ws, 2, 16)  # B
                self.ew.set_column_width(this_ws, 3, 16)  # B
            if not ready_data:
                continue
            for row_name, rd in list(ready_data.get("row_data").items()):
                output_rd = rd.values()
                self.ew.write_row_cells(this_ws, output_rd)
        self.printout_content_data(len(summary_names))
        final_excel_file = os.path.splitext(self.output_excel)[0] + ".xlsx"
        final_excel_file = xTools.get_real_path(final_excel_file, self.output)
        self.ew.save_to_excel(final_excel_file)

    def printout_content_data(self, sheet_index):
        c_s = self.ew.wb.create_sheet("Content", sheet_index)
        titles = ("id", "sheet_name", "src", "dst", "src_section", "dst_section", "src_name", "dst_name")
        title_data = list()
        for i, t in enumerate(titles):
            title_data.append(dict(cell_range=dict(row=1, column=i+1), value=t.upper(),
                                   font=self.ew.font_bold, border=self.ew.borders_lrtb["1111"]))
        id_number = 0
        column_width_dict = dict()
        item_data = list()
        for sheet_name, sheet_data in list(self.final_excel_data_dict.items()):
            if "Summary" == sheet_name or "Summary_" in sheet_name:
                continue
            row_number = id_number + 2
            item_data.append(dict(cell_range=dict(row=row_number, column=1), value=id_number+1))
            for n, t in enumerate(titles[1:]):
                if t == "sheet_name":
                    v = sheet_name
                else:
                    v = sheet_data.get(t)
                    if isinstance(v, list):
                        v = ",".join(v)
                column_width_dict.setdefault(n, 0)
                column_width_dict[n] = max(column_width_dict[n], len(v)+3, 12)
                x = dict(cell_range=dict(row=row_number, column=n+2), value=xUtils.string_2_int_float_percentage(v)[0])
                if t == "sheet_name":
                    x["hyperlink"] = "#'{}'!A1".format(v)
                    x["font"] = self.ew.font_link_content
                else:
                    x["font"] = self.ew.font_default
                x["border"] = self.ew.borders_lrtb["0000"]
                item_data.append(x)
            id_number += 1
        self.ew.write_row_cells(c_s, title_data)
        self.ew.write_row_cells(c_s, item_data)
        for k, w in list(column_width_dict.items()):
            self.ew.set_column_width(c_s, k+2, w)

    def add_row_data(self, d_id, design_name, sheet_data):
        r_data = OrderedDict()
        will_skip_this_row_when_calculate = 0
        _c = 1
        r_data["ID"] = dict(cell_range=dict(row=self.row_number, column=1), value=d_id,
                            border=self.ew.borders_lrtb["1000"], font=self.ew.font_default)
        r_data["Design"] = dict(cell_range=dict(row=self.row_number, column=2), value=design_name,
                                border=self.ew.borders_lrtb["0100"], font=self.ew.font_default)

        def __add_cell(raw_titles, raw_design_data, start_column, prefix=""):
            for ccc, bar in enumerate(raw_titles):
                if bar == "Design":
                    continue
                z = dict(cell_range=dict(row=self.row_number, column=start_column), font=self.ew.font_default)
                z_value = raw_design_data.get(bar)
                z_value, is_percentage = xUtils.string_2_int_float_percentage(z_value)
                z["value"] = z_value
                if z_value == 'NA':
                    z['alignment'] = self.ew.alignment_right
                elif is_percentage:
                    z["number_format"] = self.ew.number_percentage
                elif isinstance(z_value, float):
                    z['number_format'] = self.ew.number_float
                z_border_key = "0100" if (len(raw_titles) == ccc + 1) else "0000"
                z["border"] = self.ew.borders_lrtb[z_border_key]
                r_data[prefix+bar] = z
                start_column += 1
            return start_column

        src_titles = sheet_data.get("src_titles")
        src_design_data = sheet_data.get("src_data", dict()).get(design_name, dict())
        dst_titles = sheet_data.get("dst_titles")
        dst_design_data = sheet_data.get("dst_data", dict()).get(design_name, dict())
        _c = __add_cell(src_titles, src_design_data, 3, prefix="")
        _c = __add_cell(dst_titles, dst_design_data, _c, prefix="#")

        # use previous r_data
        compare_titles = sheet_data.get("compare_titles")
        for cc, foo in enumerate(compare_titles):
            x = dict(cell_range=dict(row=self.row_number, column=_c + cc))
            key_left, key_right = foo, "#" + foo
            position_left = get_column_letter(list(r_data.keys()).index(key_left) + 1)
            position_right = get_column_letter(list(r_data.keys()).index(key_right) + 1)
            if foo.endswith("sha1"):
                x["value"] = '=IF({0}{2}={1}{2}, "Same", "Diff")'.format(position_right, position_left, self.row_number)
            else:
                x["value"] = '=IFERROR({0}{2}/{1}{2} - 1, "-")'.format(position_right, position_left, self.row_number)
            x["number_format"] = self.ew.number_percentage
            border_key = "0100" if (len(compare_titles) == cc + 1) else "0000"
            x["border"] = self.ew.borders_lrtb[border_key]
            if foo.endswith("sha1"):
                real_value_string = r_data.get(key_right)["value"] == r_data.get(key_left)["value"]
            else:
                real_value_string = "{} / {} - 1".format(r_data.get(key_right)["value"], r_data.get(key_left)["value"])

            _real_value, _new_font, _will_skip = self.set_real_value_font(foo, real_value_string)
            if not will_skip_this_row_when_calculate:
                will_skip_this_row_when_calculate = _will_skip
            x["real_value"] = _real_value
            x["font"] = _new_font
            r_data["VS_" + foo] = x

        _c = _c + len(compare_titles)
        for cc, foo in enumerate(compare_titles):
            x = dict(cell_range=dict(row=self.row_number, column=_c + cc), font=self.ew.font_default)
            position_one = get_column_letter(list(r_data.keys()).index("VS_" + foo) + 1)
            bottom_row = 2 + len(sheet_data.get("design_names")) + 1
            x["value"] = '=IFERROR({0}{1} - {0}{2}, "-")'.format(position_one, self.row_number, bottom_row)
            x["number_format"] = self.ew.number_percentage
            border_key = "0100" if (len(compare_titles) == cc + 1) else "0000"
            x["border"] = self.ew.borders_lrtb[border_key]
            if not self.no_ad_data:
                r_data["AD_" + foo] = x
        self.row_number += 1
        if will_skip_this_row_when_calculate:
            for k, v in list(r_data.items()):
                v["fill"] = self.ew.fill_skip
        r_data[self.key_skip] = will_skip_this_row_when_calculate
        return r_data

    def set_real_value_font(self, key_name, raw_string):
        red_range = self.detail_options.get("red_{}".format(key_name),  self.detail_options.get("red_default"))
        green_range = self.detail_options.get("green_{}".format(key_name), self.detail_options.get("green_default"))
        skip_this_row = 0
        real_font = self.ew.font_default
        if isinstance(raw_string, bool):
            return str(raw_string), real_font, 0
        try:
            real_value = eval(raw_string)
            if red_range:
                judge_string = "{} {}".format(real_value, red_range)
                if eval(judge_string):
                    real_font = self.ew.font_red
                    if self.detail_options.get("skip_red") == "1":
                        skip_this_row = 1
            if green_range:
                judge_string = "{} {}".format(real_value, green_range)
                if eval(judge_string):
                    real_font = self.ew.font_green
        except (NameError, ZeroDivisionError, SyntaxError):
            if red_range or green_range:
                skip_this_row = 1
            real_value = "-"
        return real_value, real_font, skip_this_row

    def get_vs_formula_range(self, sheet_row_data):
        self.temp_failed_cases_number, self.temp_whole_cases_number, self.temp_skipped_cases_number = 0, 0, 0
        _range1 = list()
        i = 0
        for row_name, cell_data in list(sheet_row_data.items()):
            i += 1
            if not row_name.startswith("ID_"):
                continue
            self.temp_whole_cases_number += 1
            x_skip = cell_data.get(self.key_skip, 1)
            if x_skip:
                self.temp_failed_cases_number += 1
                continue
            _range1.append(i)

        # out of range
        ov = float(self.detail_options.get("out_of_range_value"))
        o_o_r_f = xTools.comma_split(self.detail_options.get("out_of_range_fields"))
        range_value = (self.temp_whole_cases_number - self.temp_failed_cases_number) * ov
        for i, (row_name, cell_data) in enumerate(list(sheet_row_data.items())):
            if not row_name.startswith("ID_"):
                continue
            for title_key, a_cell in list(cell_data.items()):
                if title_key.startswith("VS_"):
                    k_name = re.sub("VS_", "", title_key)
                    if k_name not in o_o_r_f:
                        continue
                    _rv = a_cell.get("real_value")
                    try:
                        if abs(_rv) > range_value:
                            _pos = i + 1
                            if _pos in _range1:
                                _range1.remove(_pos)
                            a_cell["fill"] = self.ew.fill_out_of_range
                            if not cell_data.get(self.key_skip):  # only count it one time
                                self.temp_skipped_cases_number += 1
                            cell_data[self.key_skip] = 1
                    except TypeError:
                        pass
        #
        _tmp, _range2 = list(), list()
        for foo in _range1:
            _tmp.append(foo)
            if (foo + 1) not in _range1:
                _range2.append((_tmp[0], _tmp[-1]))
                _tmp = list()
        excel_range = list()
        for (zuo, you) in _range2:
            # If remove $, will encounter with the Error Formula Omits Adjacent Cells
            if zuo == you:
                excel_range.append("${pos}$%d" % zuo)
            else:
                excel_range.append("${pos}$%d:${pos}$%d" % (zuo, you))
        excel_range = ",".join(excel_range)
        return excel_range

    def add_statistics_data(self, row_data, excel_range, formula_names):
        keys = list(row_data.get("row_2").keys())
        for formula_name in formula_names + ["GAP"]:
            x = OrderedDict()
            _mr = dict(start_row=self.row_number, end_row=self.row_number, start_column=1, end_column=2)
            x[formula_name] = dict(merge_range=_mr, value=formula_name + ":", font=self.ew.font_bold,
                                   border=self.ew.borders_lrtb["1111"], alignment=self.ew.alignment_right)
            if excel_range:
                if formula_name == "GAP":
                    _initial_value = "MAX({0}) - MIN({0})".format(excel_range)
                else:
                    _initial_value = "{}({})".format(formula_name, excel_range)
            else:
                _initial_value = ""

            for i, k in enumerate(keys):
                if i < 2:
                    continue
                if k.startswith("AD_") or re.search("(version|device|xTestID)", k):
                    my_value = "-"
                else:
                    if _initial_value:
                        my_value = _initial_value.format(pos=get_column_letter(i + 1))
                        my_value = '=IFERROR({}, "-")'.format(my_value)
                    else:
                        my_value = "-"
                x[k] = dict(cell_range=dict(row=self.row_number, column=i + 1), font=self.ew.font_default,
                            border=self.ew.borders_lrtb["1111"])
                if k in ("fmax_pap", "pap", "#fmax_pap", "#pap") or k.startswith("VS_"):
                    x[k]["number_format"] = self.ew.number_percentage
                else:
                    x[k]["number_format"] = self.ew.number_float
                x[k]["value"] = my_value
            row_data[formula_name] = x
            self.row_number += 1

    def add_win_tie_lose_failed_data(self, row_data, formula_range):
        win_tie_lose_failed_detail = list()
        win_tie_lose_failed_detail.append("")
        win_tie_lose_failed_detail.append(dict(nickname="K1", name="Win",   fmax=[">=0.05"], others=["<=-0.05"]))
        win_tie_lose_failed_detail.append(dict(nickname="K2", name="Tie", others=[">-0.05", "<0.05"]))
        win_tie_lose_failed_detail.append(dict(nickname="K3", name="Lose",  fmax=["<-0.05"], others=[">0.05"]))
        win_tie_lose_failed_detail.append("")
        win_tie_lose_failed_detail.append(dict(nickname="K4", name="change <= -15%",      others=["<=-0.15"]))
        win_tie_lose_failed_detail.append(dict(nickname="K5", name="-15%< change <= -5%", others=[">-0.15", "<=-0.05"]))
        win_tie_lose_failed_detail.append(dict(nickname="K6", name="-5%< change <= 5%",   others=[">-0.05", "<=0.05"]))
        win_tie_lose_failed_detail.append(dict(nickname="K7", name="5%< change <= 15%",   others=[">0.05", "<=0.15"]))
        win_tie_lose_failed_detail.append(dict(nickname="K8", name="change > 15%",        others=[">0.15"]))

        titles = list(row_data.get("row_2").keys())
        x_item_name = ""

        for foo in win_tie_lose_failed_detail:
            self.row_number += 1
            if not foo:
                continue
            new_data = dict()
            start_vs = 0
            _nick = foo.get("nickname")
            for i, t in enumerate(titles):
                if not start_vs:
                    start_vs = t.startswith("VS_")
                    if start_vs:
                        x_item_name = foo.get("name")
                        if len(x_item_name) > 10:
                            new_data[_nick] = dict(merge_range=dict(start_column=i-1, end_column=i,
                                                                    start_row=self.row_number, end_row=self.row_number))
                        else:
                            new_data[_nick] = dict(cell_range=dict(row=self.row_number, column=i))
                        new_data[_nick].update(dict(value=x_item_name,
                                                    font=self.ew.font_bold, border=self.ew.borders_lrtb["1111"]))
                if start_vs:
                    if not t.startswith("VS_"):
                        break
                    compare_fmax_list = foo.get('fmax')
                    compare_others_list = foo.get('others')
                    if not compare_fmax_list:
                        compare_list = compare_others_list
                    else:
                        compare_list = compare_fmax_list if t == "VS_fmax" else compare_others_list
                    new_data[t] = dict(cell_range=dict(column=i+1, row=self.row_number), font=self.ew.font_bold,
                                       border=self.ew.borders_lrtb["1111"])
                    original_formula = formula_range.format(pos=get_column_letter(i + 1))
                    new_data[t]["value"] = xUtils.get_range_formula(compare_list, original_formula)
            if new_data:
                row_data[x_item_name] = new_data
