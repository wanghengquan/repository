#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Initial:
  date: 3:38 PM 12/1/2021
  file: demo_excel.py
  author: Shawn Yan
Description:

"""
import json
import random
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, NamedStyle, numbers
from openpyxl.styles import fills
from openpyxl.utils import get_column_letter
from collections import OrderedDict


class WriteExcel(object):
    def __init__(self):
        self.wb = Workbook()
        self.rgb_list = ('FF0000', "00FF00", "0000FF", "3C0069", "9696F0")

    def process(self):
        first_sheet = self.wb.active
        self.write_colors()
        self.write_borders()
        self.write_font()
        self.write_fill()
        self.write_mixed()
        self.wb.remove(first_sheet)
        self.wb.save("excel_demo.xlsx")

    def get_random_rgb(self):
        return random.choice(self.rgb_list)

    def write_colors(self):
        _ws = self.wb.create_sheet("Colors")
        _row = 1
        _col = 2
        step = 2 ** 4 - 1
        max_value = 0xFF + step
        for red in range(0, max_value, step):
            _row += 2
            _col = 2
            for green in range(0, max_value, step):
                for blue in range(0, max_value, step):
                    color_code = "{:02x}{:02x}{:02x}".format(red, green, blue)
                    color_code = color_code.upper()
                    _this_cell = _ws.cell(row=_row, column=_col, value='#{}'.format(color_code))
                    another_cell = _ws.cell(row=_row, column=_col + 1)
                    another_cell.fill = PatternFill(fill_type='solid', start_color=color_code, end_color=color_code)
                    _col += 2
                    if _col > 20:
                        _row += 1
                        _col = 2

    def write_borders(self):
        _ws = self.wb.create_sheet("Borders")
        side_styles = (None, 'dashDot', 'dashDotDot', 'dashed', 'dotted',
                       'double', 'hair', 'medium', 'mediumDashDot', 'mediumDashDotDot',
                       'mediumDashed', 'slantDashDot', 'thick', 'thin')
        borders = OrderedDict()
        for _s in set(side_styles):
            for _c in set(self.rgb_list):
                key = "style: {}; color: {}".format(_s, _c)
                this_side = Side(style=_s, color=_c)
                this_border = Border(left=this_side, right=this_side, top=this_side, bottom=this_side)
                borders[key] = this_border
        _ws.column_dimensions[get_column_letter(2)].width = 40
        _ws.column_dimensions[get_column_letter(4)].width = 40
        _row_tight = 2
        _row_loose = 2
        for k, v in list(borders.items()):
            cell_tight = _ws.cell(_row_tight, 2, value=k)
            cell_tight.border = v
            _row_tight += 1

            cell_loose = _ws.cell(_row_loose, 4, value=k)
            cell_loose.border = v
            _row_loose += 2

    def write_font(self):
        _ws = self.wb.create_sheet("Font", 0)
        settings = dict(
            name=("Arial", "Georgia"),
            color=("000000", "005555", "FF0000"),
            size=(12, 45),
            bold=(True, False),
            italic=(True, False),
            strikethrough=(True, False),
            underline=('single', 'double', 'singleAccounting', 'doubleAccounting'),
            vertAlign=('baseline', 'superscript', 'subscript')
        )
        _row = 0
        for k, v in list(settings.items()):
            _row += 1
            for _c, vv in enumerate(v):
                this_cell = _ws.cell(row=_row, column=_c + 1)
                this_cell.value = "{} = {}".format(k, vv)
                _font_dict = {k: vv}
                this_cell.font = Font(**_font_dict)
        for i in range(_ws.max_column):
            _ws.column_dimensions[get_column_letter(i + 1)].width = 30

    def get_fill_dict(self, type_name):
        x = dict(fill_type=type_name,
                 start_color=self.get_random_rgb(),
                 end_color=self.get_random_rgb()
                 )
        return x

    def write_fill(self):
        _ws = self.wb.create_sheet("fill")
        _ws.column_dimensions[get_column_letter(2)].width = 80
        _ws.column_dimensions[get_column_letter(4)].width = 80

        _row = 0
        _ws.cell(row=1, column=1, value='PatternFill')
        for _row_minus_1, foo in enumerate(fills.fills):
            _row += 2
            for _column in (2, 4):
                fill_dict = self.get_fill_dict(foo)
                this_cell = _ws.cell(row=_row, column=_column, value=json.dumps(fill_dict))
                this_cell.fill = PatternFill(**fill_dict)

        _ws.cell(row=_row + 1, column=1, value='GradientFill')
        for gf_type in ("linear", "path"):
            for i in range(15, 360, 23):
                _row += 2
                gf_dict = dict(type=gf_type,
                               degree=i,
                               stop=(self.get_random_rgb(), self.get_random_rgb()))
                this_cell = _ws.cell(row=_row, column=2, value=json.dumps(gf_dict))
                this_cell.fill = GradientFill(**gf_dict)
                _ws.row_dimensions[_row].height = 46
            for j in range(20):
                _row += 2
                gf_dict = dict(type=gf_type,
                               left=random.randint(1, 10) / 10,
                               top=random.randint(1, 10) / 10,
                               stop=(self.get_random_rgb(), self.get_random_rgb()))
                gf_dict_2 = dict(type=gf_type, left=0.5, top=0.5, stop=gf_dict.get("stop"))
                this_cell = _ws.cell(row=_row, column=2, value=json.dumps(gf_dict))
                this_cell.fill = GradientFill(**gf_dict)
                this_cell = _ws.cell(row=_row, column=4, value=json.dumps(gf_dict_2))
                this_cell.fill = GradientFill(**gf_dict_2)
                _ws.row_dimensions[_row].height = 46

    def write_mixed(self):
        # create a sheet named Mixed and put it at the first sheet
        _ws = self.wb.create_sheet("Mixed", 0)
        # set column width
        _ws.column_dimensions[get_column_letter(2)].width = 80
        # select a cell, row/column number starts from 1.
        one_cell = _ws.cell(row=1, column=2)
        one_cell.value = "Design"
        # select another cell, row/column number starts from 1.
        two_cell = _ws["C2"]
        two_cell.value = "Version"
        # write a list. It's better to use float/integer value in the list.
        _ws.append(["one", "two", "three", "4", 5, "6.2", 7.8, "=SUM(D3:G3)", "=B3", "=fill!D20"])
        # add hyperlinks
        link_font = Font(underline='single', color='0563C1')
        _baidu = '=HYPERLINK("http://www.baidu.com", "Baidu")'
        three_cell = _ws.cell(row=4, column=3)
        three_cell.value = _baidu
        three_cell.font = link_font
        four_cell = _ws.cell(row=4, column=5, value="toBorders")
        four_cell.hyperlink = "#Borders!A1"
        four_cell.font = link_font
        # customer style
        thin = Side(border_style="thin", color="00FE0C")
        double = Side(border_style="double", color="ff0000")
        ns_1 = dict(name="NamedStyle1",
                    font=Font(b=True, color="FF0000"),
                    fill=GradientFill(degree=20, stop=("000000", "FFF0FF")),
                    border=Border(top=double, left=thin, right=thin, bottom=double),
                    alignment=Alignment(horizontal="right", vertical="center"))
        style_one = NamedStyle(**ns_1)
        _ws.row_dimensions[6].height = 46
        cell_a = _ws.cell(row=6, column=2, value="use customer style: NamedStyle1")
        cell_a.style = style_one
        _ws.row_dimensions[8].height = 46
        cell_b = _ws.cell(row=8, column=2, value="use original style: Bad")
        cell_b.style = "Bad"
        cell_c = _ws.cell(row=10, column=2, value="use original style: Linked Cell")
        cell_c.style = "Linked Cell"

        row = 12
        for k, fm in list(numbers.BUILTIN_FORMATS.items()):
            num_cell = _ws.cell(row=row, column=1, value=3.1415926)
            _ws.cell(row=row, column=2, value="string: {}".format(fm))
            num_cell.number_format = fm
            row += 1


if __name__ == "__main__":
    we = WriteExcel()
    we.process()
