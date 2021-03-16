#!/usr/bin/python
# -*- coding: utf-8 -*-
"""
Initial:
  date: 10:36 AM 10/23/2019
  file: auto_regression.py
  author: Shawn Yan
Description:
  >> python auto_regression.py --dump-only
  >> python auto_regression.py --dump-done
  >> python auto_regression.py --gen-excel-only
  >> python auto_regression.py

"""
import os
import re
import argparse
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from LoppLagoonLake import tools
from LoppLagoonLake import log4u

#################
EXCEL_TITLES = ["Public Case Info", "FPGA Extra INFO"]
EXCEL_WIDTH = [21, 29]
EXCEL_FIELDS = ["Order", "NoUse", "Title", "Section", "design_name", "TestLevel",
                "TestScenarios", "Description", "Type", "Priority", "Automated",
                "CaseInfo", "Environment", "LaunchCommand", "Software", "System",
                "Machine", "Sorting", "CRs", "Author", "Create", "Update", "Family", "Slice",
                "PIO", "DSP", "EBR", "RTL", "TestBench", "Flow"]
COLUMN_WIDTH_DICT = dict(Title=25, Section=18, design_name=25, CaseInfo=30,
                         Environment=25, LaunchCommand=30, Software=22)
SUITE_INFO = [
    "[suite_info]",
    ("project_id", 8),
    ("suite_name", "CR-Regression"),
    ("CaseInfo", "repository={svn_path}"),
    ("", "suite_path={suite_name}"),
    ("Environment", ""),
    ("LaunchCommand", "cmd={base_cmd}"),
    ("Software", "{base_sw}"),
    ("System", ""),
    ("Machine", "{base_machine}"),
    "",
    "",
    "END"
]
#################


def get_value(key, a_dict, default_value):
    value = a_dict.get(key)
    if not value:
        value = default_value
    return value


def get_value_with_comma(a_dict):
    _v = list()
    for k, v in list(a_dict.items()):
        if not v:
            continue
        _v.append("{} = {}".format(k, v))
    return ";".join(_v)


class AutoRegression(object):
    def __init__(self):
        self.info_name = "run_info.ini"  # previous is .txt
        self.default_svn = "http://lsh-tmp/radiant/trunk/customer_cr"
        self.diamond_options_json_file = "diamond.json"
        self.radiant_options_json_file = "radiant.json"
        self.diamond_suite_excel_file = "diamond.xlsx"
        self.radiant_suite_excel_file = "radiant.xlsx"
        self.p_dev_py = re.compile(r"DEV.+run_\S+\.py")

    def process(self):
        self.get_options()
        if not self.dump_done:
            self.dump_options()
        if self.dump_only:
            return
        self.generate_excel_file()
        if self.gen_excel_only:
            return

    def get_options(self):
        hlp = "dump info options already done, use dump file straightly"

        parser = argparse.ArgumentParser()
        parser.add_argument("--svn", help="specify CR cases svn path")
        parser.add_argument("--dump-only", action="store_true", help="dump info options only")
        parser.add_argument("--dump-done", action="store_true", help=hlp)
        parser.add_argument("--gen-excel-only", action="store_true", help="generate Suite Excel file only")
        parser.add_argument("--debug", action="store_true", help="print debug message")
        parser.set_defaults(svn=self.default_svn)
        opts = parser.parse_args()
        self.svn = opts.svn
        self.dump_only = opts.dump_only
        self.dump_done = opts.dump_done
        self.gen_excel_only = opts.gen_excel_only
        self.debug = opts.debug

    def dump_options(self):
        log4u.say_it(">> Dump Options:")
        self.design_paths = list()
        log4u.say_it("  >> Get design path list ...")
        self.get_design_paths(self.svn)
        log4u.say_it(self.design_paths, "Design Paths:", show=self.debug)
        log4u.say_it("  >> dump options to json file ...")
        self.dump_to_json_file()

    def get_design_paths(self, svn_url):
        svn_list = "svn list {}".format(svn_url)
        sts, res = tools.get_status_output(svn_list)
        if sts:
            log4u.say_it("Failed to run {}".format(svn_list))
            tools.eexit()
        res_list = re.split("\n", res)
        for foo in res_list:
            if foo == self.info_name:
                self.design_paths.append(svn_url)
                break
        else:
            for foo in res_list:
                if foo.endswith("/"):
                    new_svn_url = "{}/{}".format(svn_url, foo.strip("/"))
                    self.get_design_paths(new_svn_url)

    def dump_to_json_file(self):
        for foo in (self.radiant_options_json_file, self.diamond_options_json_file):
            if os.path.isdir(foo):
                os.remove(foo)
        json_data = dict()
        json_data[self.radiant_options_json_file] = dict()
        json_data[self.diamond_options_json_file] = dict()
        for design_path in self.design_paths:
            log4u.say_it("  .. Processing {}".format(design_path))
            cat_cmd = "svn cat {}/{}".format(design_path, self.info_name)
            sts, res = tools.get_status_output(cat_cmd)
            options = tools.get_string_options("{}".format(res))
            base_dict = options.get("Base")
            software_dict = options.get("Software")
            if "radiant" in base_dict.get("test_software", "") or software_dict.get("radiant"):
                json_file = self.radiant_options_json_file
            else:
                json_file = self.diamond_options_json_file
            json_data[json_file][design_path] = options
        for k, v in list(json_data.items()):
            if v:
                json.dump(v, open(k, "w"), indent=2)

    def generate_excel_file(self):
        log4u.say_it("  Generate Suite Excel file ...")
        self.generate_suite_excel_file("diamond", self.diamond_options_json_file, self.diamond_suite_excel_file)
        self.generate_suite_excel_file("radiant", self.radiant_options_json_file, self.radiant_suite_excel_file)

    def generate_suite_excel_file(self, vendor, json_file, suite_excel_file):
        if not os.path.isfile(json_file):
            return
        self.vendor = vendor
        options = json.load(open(json_file))
        wb = Workbook()
        suite_sheet = wb.create_sheet("suite", 0)
        case_sheet = wb.create_sheet("case", 1)
        wb.remove(wb["Sheet"])
        #######
        self._left = Alignment(horizontal="left")
        self._center = Alignment(horizontal="center")
        self.write_suite_sheet(suite_sheet, vendor, options)
        self.write_case_sheet(case_sheet, vendor, options)
        #######
        wb.save(suite_excel_file)

    def write_suite_sheet(self, sheet, vendor, options):
        suite_dict = dict()
        suite_dict["svn_path"] = os.path.dirname(self.svn)
        suite_dict["suite_name"] = os.path.basename(self.svn)
        suite_dict["base_cmd"] = "python DEV/bin/run_{}.py".format(vendor)
        machine_list = list()
        for k, v in list(options.items()):
            z = v.get("Machine", dict()).get("group")
            machine_list.append(z)
        machine_set = set(machine_list)
        cnt_number = 0
        for foo in machine_set:
            this_cnt = machine_list.count(foo)
            if this_cnt > cnt_number:
                cnt_number = this_cnt
                self.pub_group = foo
        suite_dict["base_sw"] = "{} = default_tool_name".format(vendor)
        suite_dict["base_machine"] = "group = {}".format(self.pub_group)
        for i, bar in enumerate(SUITE_INFO):
            if isinstance(bar, str):
                sheet.cell(row=i + 1, column=1, value=bar)
            else:
                sheet.cell(row=i + 1, column=1, value=bar[0])
                if isinstance(bar[1], int):
                    v = bar[1]
                else:
                    v = bar[1].format(**suite_dict)
                sheet.cell(row=i + 1, column=2, value=v).alignment = self._left
        sheet.column_dimensions[get_column_letter(1)].width = 20
        sheet.column_dimensions[get_column_letter(2)].width = 38

    def get_command(self, cmd_dict):
        _command = list()
        for a, b in list(cmd_dict.items()):
            if not b:
                continue
            if a == "cmd":
                b_list = b.split()
                for timon in b_list:
                    if self.p_dev_py.search(timon):
                        if "runSquish" in b_list[1]:
                            _command.append("cmd = {}".format(" ".join(b_list[:])))
                        else:
                            _command.append("cmd = {}".format(" ".join(b_list[2:])))
                        break
                else:
                    _command.append("{} = {}".format(a, b))
            else:
                _command.append("{} = {}".format(a, b))
        return ";".join(_command)

    @staticmethod
    def get_sw(vendor, sw_dict):
        _sw = list()
        for a, b in list(sw_dict.items()):
            if not b:
                continue
            if a == vendor:
                continue
            _sw.append("{} = {}".format(a, b))
        return ";".join(_sw)

    def get_machine(self, machine_dict):
        _machine = list()
        for a, b in list(machine_dict.items()):
            if not b:
                continue
            if a == "group":
                if b == self.pub_group:
                    continue
            _machine.append("{} = {}".format(a, b))
        return ";".join(_machine)

    def write_case_sheet(self, sheet, vendor, options):
        keys = list(options.keys())
        keys.sort()
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=EXCEL_WIDTH[0])
        sheet.merge_cells(start_row=1, start_column=EXCEL_WIDTH[0] + 1, end_row=1, end_column=EXCEL_WIDTH[1])
        sheet.cell(row=1, column=1, value=EXCEL_TITLES[0]).alignment = self._center
        sheet.cell(row=1, column=EXCEL_WIDTH[0] + 1, value=EXCEL_TITLES[1]).alignment = self._center
        for i, field_name in enumerate(EXCEL_FIELDS):
            sheet.cell(row=2, column=i+1, value=field_name)
            _wid = COLUMN_WIDTH_DICT.get(field_name)
            if _wid:
                sheet.column_dimensions[get_column_letter(i+1)].width = _wid
            sheet.auto_filter.ref = "A2:{}{}".format(get_column_letter(len(EXCEL_FIELDS)), len(options)+8)
        for i, k in enumerate(keys):
            v = options.get(k)
            machine_dict = v.get("Machine", dict())
            case_info_dict = v.get("CaseInfo", dict())
            system_dict = v.get("System", dict())
            base_dict = v.get("Base", dict())
            general_dict = v.get("General", dict())
            env_dict = v.get("Environment", dict())
            cmd_dict = v.get("LaunchCommand", dict())
            sw_dict = v.get("Software", dict())
            case_data = dict()
            case_data["Order"] = i + 1
            case_data["NoUse"] = get_value("nouse", general_dict, "")
            _title = re.sub(self.svn, "", k).strip("/")
            case_data["Title"] = get_value("title", general_dict, _title)
            case_data["Section"] = get_value("section", general_dict, "{}_CR".format(vendor))
            case_data["design_name"] = get_value("design_name", general_dict, _title)
            case_data["TestLevel"] = int(get_value("testlevel", general_dict, 1))
            case_data["TestScenarios"] = get_value("testscenarios", general_dict, "")
            case_data["Description"] = get_value("description", general_dict, "")
            case_data["Type"] = get_value("type", general_dict, "")
            case_data["Priority"] = get_value("priority", general_dict, "")
            case_data["Automated"] = get_value("automated", general_dict, "")
            case_data["CaseInfo"] = get_value_with_comma(case_info_dict)
            case_data["Environment"] = get_value_with_comma(env_dict)
            case_data["LaunchCommand"] = self.get_command(cmd_dict)
            case_data["Software"] = self.get_sw(vendor, sw_dict)
            case_data["System"] = get_value_with_comma(system_dict)
            case_data["Machine"] = self.get_machine(machine_dict)
            case_data["CRs"] = get_value("crs", general_dict, "")
            case_data["Author"] = get_value("author", base_dict, "")
            for col, bar in enumerate(EXCEL_FIELDS):
                sheet.cell(row=i+3, column=col + 1, value=case_data.get(bar, ""))


if __name__ == "__main__":
    p = AutoRegression()
    p.process()
